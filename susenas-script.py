# ============================================================
# Poverty Profile Analysis - Python conversion of SPSS syntax
# Source: 72_ssn_202403 dataset (Susenas March 2024)
# Output: Excel workbook with all tables (1-69)
# ============================================================

import warnings

import numpy as np
import pandas as pd
import simpledbf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ============================================================
# CONFIG - Update these paths to match your environment
# ============================================================
PATH_KP43 = "data/2026/1707-ssn202503-kp-blok43.dbf"
PATH_IND_1 = "data/2026/1707-ssn202503-kor-ind-1.dbf"
PATH_IND_2 = "data/2026/1707-ssn202503-kor-ind-2.dbf"
PATH_RT = "data/2026/1707-ssn202503-kor-rt.dbf"
PATH_KP41 = "data/2026/1707-ssn202503-kp-blok41.dbf"
PATH_KP42 = "data/2026/1707-ssn202503-kp-blok42.dbf"
OUTPUT = "output/Poverty_Profile_Tables.xlsx"

# Poverty parameters for r102=7 (Lebong)
FK_VALUE = 0.8135755811606
GK_VALUE = 540904
REGION_CODE = 7

# ============================================================
# HELPER FUNCTIONS
# ============================================================


def weighted_mean(values, weights):
    mask = values.notna() & weights.notna()
    if mask.sum() == 0:
        return np.nan
    return np.average(values[mask], weights=weights[mask])


def weighted_count(values, weights):
    mask = values.notna() & weights.notna()
    return weights[mask].sum()


def weighted_pct_col(df, row_var, col_var, weight_var):
    """Weighted column percentage cross-tabulation."""
    result = {}
    for col_val in sorted(df[col_var].dropna().unique()):
        sub = df[df[col_var] == col_val]
        counts = sub.groupby(row_var)[weight_var].sum()
        total = counts.sum()
        pcts = (counts / total * 100).round(2) if total > 0 else counts * 0
        result[col_val] = pd.DataFrame({"N": counts.round(0), "%": pcts})
    return result


def weighted_mean_by(df, value_var, by_vars, weight_var):
    """Weighted mean grouped by variables."""
    rows = []
    for keys, grp in df.groupby(by_vars):
        val = weighted_mean(grp[value_var], grp[weight_var])
        row = dict(zip(by_vars, keys if isinstance(keys, tuple) else [keys]))
        row["mean"] = round(val, 3) if pd.notna(val) else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


def style_header(cell, bg="1F4E79"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_subheader(cell):
    cell.font = Font(bold=True, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color="BDD7EE")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data(cell, bold=False):
    cell.font = Font(bold=bold, name="Arial", size=9)
    cell.alignment = Alignment(horizontal="right", vertical="center")


def style_label(cell, bold=False):
    cell.font = Font(bold=bold, name="Arial", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def calc_pct_ensure_100(df, n_cols, pct_cols, row_order, total_label="Total"):
    """
    Calculate percentages ensuring they sum to exactly 100.
    - df: DataFrame with index including row_order and total_label
    - n_cols: list of N column names (e.g., ["N mkako=0", "N mkako=1"])
    - pct_cols: list of % column names (e.g., ["% mkako=0", "% mkako=1"])
    - row_order: list of row labels (excluding Total)
    - total_label: the total row label
    """
    for n_col, pct_col in zip(n_cols, pct_cols):
        total_n = df.loc[total_label, n_col]
        if total_n > 0:
            pct_sum = 0
            # Calculate for all except last
            for idx in row_order[:-1]:
                pct = round(df.loc[idx, n_col] / total_n * 100, 2)
                df.loc[idx, pct_col] = pct
                pct_sum += pct
            # Last item: ensure sum = 100
            df.loc[row_order[-1], pct_col] = round(100 - pct_sum, 2)
            # Total row
            df.loc[total_label, pct_col] = 100.00


def write_table(ws, title, df, start_row=1):
    """Generic table writer to worksheet."""
    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=len(df.columns) + 1,
    )
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, name="Arial", size=11)
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.fill = PatternFill("solid", start_color="1F4E79")
    title_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    # Header row
    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="Kategori").font = Font(
        bold=True, name="Arial", size=10
    )
    for ci, col in enumerate(df.columns, start=2):
        c = ws.cell(row=header_row, column=ci, value=str(col))
        style_header(c)

    # Data rows
    for ri, (idx, row) in enumerate(df.iterrows(), start=header_row + 1):
        lc = ws.cell(row=ri, column=1, value=str(idx))
        style_label(lc)
        for ci, val in enumerate(row, start=2):
            dc = ws.cell(row=ri, column=ci, value=val)
            style_data(dc)
        if str(idx).lower() in ["total", "jumlah"]:
            for ci in range(1, len(df.columns) + 2):
                ws.cell(row=ri, column=ci).font = Font(bold=True, name="Arial", size=9)

    # Column widths
    ws.column_dimensions["A"].width = 40
    for ci in range(2, len(df.columns) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    return ri + 2  # next available row


# ============================================================
# LOAD DATA
# ============================================================

print("Loading datasets...")
df_kp43 = simpledbf.Dbf5(PATH_KP43).to_dataframe()
df_ind_1 = simpledbf.Dbf5(PATH_IND_1).to_dataframe()
df_ind_2 = simpledbf.Dbf5(PATH_IND_2).to_dataframe()
df_rt = simpledbf.Dbf5(PATH_RT).to_dataframe()

# Lowercase all column names first
for df in [df_kp43, df_ind_1, df_ind_2, df_rt]:
    df.columns = df.columns.str.lower()

# Merge IND files horizontally (split due to DBF column limit)
# Drop duplicate common columns from IND-2 except merge keys
merge_keys = ["urut", "r401"]
common_cols = ["psu", "ssu", "strata", "r101", "r102", "r105", "fwt"]
df_ind_2_unique = df_ind_2.drop(
    columns=[c for c in common_cols if c in df_ind_2.columns]
)
df_ind = pd.merge(df_ind_1, df_ind_2_unique, on=merge_keys, how="inner")

# ============================================================
# SECTION 1: POVERTY VARIABLES (kp43 file)
# ============================================================

print("Computing poverty variables...")

# Correction factor and poverty line
df_kp43["fk"] = np.where(df_kp43["r102"] == REGION_CODE, FK_VALUE, 0)
df_kp43["gkkako"] = np.where(df_kp43["r102"] == REGION_CODE, GK_VALUE, 0)

# Per capita expenditure with correction factor
df_kp43["kapitafk"] = df_kp43["kapita"] * df_kp43["fk"]

# Poor household flag
df_kp43["mkako"] = np.where(df_kp43["kapitafk"] < df_kp43["gkkako"], 1, 0)

# Poverty depth index P1
df_kp43["p1kako"] = np.where(
    df_kp43["mkako"] == 1,
    ((df_kp43["gkkako"] - df_kp43["kapitafk"]) / df_kp43["gkkako"]) * 100,
    0,
)

# Poverty severity index P2
df_kp43["p2kako"] = np.where(
    df_kp43["mkako"] == 1,
    ((df_kp43["gkkako"] - df_kp43["kapitafk"]) / df_kp43["gkkako"]) ** 2 * 100,
    0,
)

# Decile ranking (Ntile 10)
df_kp43["nkapita"] = (
    pd.qcut(df_kp43["kapita"], q=10, labels=False, duplicates="drop") + 1
)

# World Bank 40-40-20 grouping
df_kp43["nkapitakakowb"] = pd.cut(
    df_kp43["nkapita"],
    bins=[0, 4, 8, 10],
    labels=["1 40% terbawah", "2 40% menengah", "3 20% teratas"],
)

# Food and non-food per capita
df_kp43["foodkapita"] = df_kp43["food"] / df_kp43["r301"]
df_kp43["nonfoodkapita"] = df_kp43["nonfood"] / df_kp43["r301"]
df_kp43["sharefoodkapita"] = df_kp43["foodkapita"] / df_kp43["kapita"] * 100
df_kp43["sharenonfoodkapita"] = df_kp43["nonfoodkapita"] / df_kp43["kapita"] * 100

# Poverty dummy (percentage)
df_kp43["dmkako"] = np.where(df_kp43["mkako"] == 1, 100, 0)

# ============================================================
# MERGE POVERTY STATUS TO IND AND RT FILES
# ============================================================

print("Merging poverty status to individual and household files...")

# Columns to merge from kp43
poverty_cols = [
    "urut",
    "mkako",
    "dmkako",
    "p1kako",
    "p2kako",
    "nkapita",
    "kapita",
    "kapitafk",
    "gkkako",
    "weind",
    "foodkapita",
    "nonfoodkapita",
    "sharefoodkapita",
    "sharenonfoodkapita",
]

# Merge to df_ind (individuals)
df_ind = pd.merge(df_ind, df_kp43[poverty_cols], on="urut", how="left")

# Merge to df_rt (households)
df_rt = pd.merge(df_rt, df_kp43[poverty_cols], on="urut", how="left")

# ============================================================
# SECTION 2: INDIVIDUAL VARIABLES (ind file)
# ============================================================

print("Computing individual variables...")

# Age groups
df_ind["kelum1"] = pd.cut(
    df_ind["r407"], bins=[-1, 14, 64, 999], labels=["0-14", "15-64", "65+"]
)
df_ind["kelum2"] = pd.cut(df_ind["r407"], bins=[9, 18, 999], labels=["10-18", ">18"])
df_ind["kelum3"] = pd.cut(
    df_ind["r407"],
    bins=[14, 24, 44, 64, 999],
    labels=["15-24", "25-44", "45-64", "65+"],
)
df_ind["kelum4"] = pd.cut(
    df_ind["r407"], bins=[6, 12, 15, 18], labels=["7-12", "13-15", "16-18"]
)


# Education level (using r615 - Ijazah/STTB Tertinggi Yang Dimiliki)
# Grouping: (0,25)=Tidak Tamat SD, (1-5)=SD Sederajat, (6-10)=SMP Sederajat,
#           (11-17)=SMU Sederajat, (18-24)=Perguruan Tinggi
# For table: SD-SMP Sederajat = SD + SMP combined
def recode_ijasah(x):
    if x == 25:  # Tidak Punya Ijazah SD
        return "Tidak Tamat SD/Tidak atau belum pernah bersekolah"
    elif 1 <= x <= 10:  # SD Sederajat (1-5) + SMP Sederajat (6-10)
        return "SD-SMP Sederajat"
    elif 11 <= x <= 17:  # SMU Sederajat
        return "SMU Sederajat"
    elif 18 <= x <= 24:  # Perguruan Tinggi
        return "Perguruan Tinggi Sederajat"
    return np.nan


df_ind["kelijasah"] = df_ind["r615"].apply(recode_ijasah)

# Education order for tables
educ_order = [
    "Tidak Tamat SD/Tidak atau belum pernah bersekolah",
    "SD-SMP Sederajat",
    "SMU Sederajat",
    "Perguruan Tinggi Sederajat",
]

# Education order for tables
educ_order = [
    "Tidak Tamat SD/Tidak atau belum pernah bersekolah",
    "SD-SMP Sederajat",
    "SMU Sederajat",
    "Perguruan Tinggi Sederajat",
]

# School participation (aps) - using r611 = 2 (Masih bersekolah)
df_ind["aps"] = np.where(df_ind["r611"] == 2, 100, 0)

# Literacy (amh) - moved from r607/r608/catatar609 to r608/r609/r610
df_ind["amh"] = np.where(
    (df_ind["r608"] == 1) | (df_ind["r609"] == 1) | (df_ind["r610"] == 1),
    100,
    np.where(
        (df_ind["r608"] == 5) & (df_ind["r609"] == 5) & (df_ind["r610"] == 5), 0, np.nan
    ),
)

# Work status
# r705/r706 = NaN means "Tidak Bekerja" (not working)
df_ind["tkerja"] = np.where(df_ind["r705"].isna(), 100, 0)
df_ind["statuskerja"] = pd.cut(
    df_ind["r706"], bins=[0, 2, 6], labels=["Informal", "Formal"]
)
df_ind.loc[df_ind["r706"].isin([3, 4]), "statuskerja"] = "Formal"
df_ind.loc[df_ind["r706"].isin([1, 2, 5, 6]), "statuskerja"] = "Informal"
df_ind["kformal"] = np.where(df_ind["statuskerja"] == "Formal", 100, 0)
df_ind["kinformal"] = np.where(df_ind["statuskerja"] == "Informal", 100, 0)

# Sector
# r705: 1-6 = Pertanian, 7+ = Non Pertanian
df_ind["sektorkerja"] = pd.cut(
    df_ind["r705"], bins=[0, 6, 99], labels=["Pertanian", "Non Pertanian"]
)
df_ind["ktani"] = np.where(df_ind["sektorkerja"] == "Pertanian", 100, 0)
df_ind["kntani"] = np.where(df_ind["sektorkerja"] == "Non Pertanian", 100, 0)

# JKN ownership
jkn_cols = ["r1101_a", "r1101_b", "r1101_c", "r1101_d", "r1101_e"]
jkn_vals = ["A", "B", "C", "D", "E"]
jkn_mask = False
for col, val in zip(jkn_cols, jkn_vals):
    if col in df_ind.columns:
        jkn_mask = jkn_mask | (df_ind[col].astype(str).str.strip() == val)
df_ind["milikjkn"] = np.where(jkn_mask, 1, 0)
if "r1101_x" in df_ind.columns:
    df_ind.loc[df_ind["r1101_x"].astype(str).str.strip() == "X", "milikjkn"] = 0


# Smoking
def recode_rokok(x):
    if x == 1:
        return "Ya, setiap hari"
    elif x == 2:
        return "Ya, tidak setiap hari"
    else:
        return "Tidak/Tidak Tahu"


df_ind["rokok"] = df_ind["r1207"].apply(recode_rokok)


# Household size groups
def recode_jart(x):
    if 1 <= x <= 3:
        return "1-3"
    elif 4 <= x <= 6:
        return "4-6"
    elif x >= 7:
        return ">=7"
    return np.nan


df_ind["keljart"] = df_ind["r301"].apply(recode_jart)

# ============================================================
# SECTION 3: HOUSEHOLD VARIABLES (rt file)
# ============================================================

print("Computing household variables...")

# Proper sanitation (moved from r1809 to r1609)
df_rt["sal"] = 0
mask1 = (df_rt["r1609a"] <= 3) & (df_rt["r1609b"] == 1) & (df_rt["r1609c"] <= 2)
mask2 = (
    (df_rt["r1609a"] <= 3)
    & (df_rt["r1609b"] == 1)
    & (df_rt["r1609c"] == 4)
    & (df_rt["r105"] == 2)
)
df_rt.loc[mask1 | mask2, "sal"] = 100


df_rt["airmlayak"] = 0
good_src = {3, 4, 5, 7, 10}
bad_src = {1, 2}
other_src = {6, 8, 9, 11}

mask = (df_rt["r1610a"].isin(good_src)) & (df_rt["r1614a"].isin(good_src))
df_rt.loc[mask, "airmlayak"] = 100

mask = (df_rt["r1610a"].isin(good_src)) & (df_rt["r1614a"].isin(bad_src))
df_rt.loc[mask, "airmlayak"] = 100

mask = (df_rt["r1610a"].isin(good_src)) & (df_rt["r1614a"].isin(other_src))
df_rt.loc[mask, "airmlayak"] = 100

mask = (df_rt["r1610a"].isin(bad_src)) & (df_rt["r1614a"].isin(good_src))
df_rt.loc[mask, "airmlayak"] = 100

df_rt["sab"] = 0
df_rt.loc[df_rt["r1610a"].isin([1, 2, 3]), "sab"] = 100
df_rt.loc[(df_rt["r1610a"].isin([4, 5, 7])) & (df_rt["r1610c"] == 2), "sab"] = 100

# Floor area per capita (moved from r1804 to r1604)
df_rt["lkapita"] = df_rt["r1604"] / df_rt["r301"]
df_rt["klkapita"] = pd.cut(
    df_rt["lkapita"], bins=[-1, 7.2, 9999], labels=["<=7,2 m2", ">7,2 m2"]
)


# Roof type (moved from r1806a to r1606)
def recode_atap(x):
    if x in [1, 2]:
        return "Beton/Genteng"
    elif x == 3:
        return "Seng"
    elif x == 4:
        return "Asbes"
    else:
        return "Bambu/kayu/jerami/lainnya"


df_rt["katap"] = df_rt["r1606"].apply(recode_atap)


# Wall type (moved from r1807 to r1607)
def recode_dinding(x):
    if x == 1:
        return "Tembok"
    elif x == 2:
        return "Plesteran Anyaman Bambu/Kawat"
    elif x == 3:
        return "Kayu/Papan"
    else:
        return "Lainnya"


df_rt["kdinding"] = df_rt["r1607"].apply(recode_dinding)


# Floor type (moved from r1808 to r1608)
def recode_lantai(x):
    if 1 <= x <= 3:
        return "Marmer/Granit/Keramik/Parket/Vinil/Karpet"
    elif x == 4:
        return "Ubin/Tegel/Teraso"
    elif x == 5:
        return "Kayu/Papan"
    else:
        return "Semen/Bata/Bambu/Tanah/Lainnya"


df_rt["klantai"] = df_rt["r1608"].apply(recode_lantai)


# Electricity (moved from r1816 to r1616)
def recode_listrik(x):
    if x in [1, 2]:
        return "Listrik PLN"
    elif x == 3:
        return "Listrik Non PLN"
    elif x == 4:
        return "Bukan listrik"
    return np.nan


df_rt["klistrik"] = df_rt["r1616"].apply(recode_listrik)

# ============================================================
# BUILD EXCEL WORKBOOK
# ============================================================

print("Writing Excel tables...")
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet


# ---- Utility: create or get sheet ----
def get_ws(name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


# ============================================================
# TABLE 0: POVERTY SUMMARY (P0, P1, P2)
# ============================================================

ws = get_ws("T0_Poverty_Summary")
ws.append(["Indikator", "Nilai"])
style_header(ws["A1"])
style_header(ws["B1"])

p0 = weighted_mean(df_kp43["dmkako"], df_kp43["weind"])
p1 = weighted_mean(df_kp43["p1kako"], df_kp43["weind"])
p2 = weighted_mean(df_kp43["p2kako"], df_kp43["weind"])
jml_miskin = df_kp43.loc[df_kp43["mkako"] == 1, "weind"].sum()

rows = [
    ["Persentase Penduduk Miskin (P0) %", round(p0, 2)],
    ["Indeks Kedalaman Kemiskinan (P1)", round(p1, 3)],
    ["Indeks Keparahan Kemiskinan (P2)", round(p2, 3)],
    ["Jumlah Penduduk Miskin (jiwa)", round(jml_miskin, 0)],
    ["Garis Kemiskinan (Rp/kapita/bulan)", GK_VALUE],
    ["Faktor Koreksi", FK_VALUE],
]
for r in rows:
    ws.append(r)
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 20

# ============================================================
# TABLE 1: P0, P1, P2 by r102
# ============================================================

ws = get_ws("T1_P0_P1_P2")
rows_t1 = []
for r102_val, grp in df_kp43.groupby("r102"):
    n_total = grp["weind"].sum()
    n_miskin = grp.loc[grp["mkako"] == 1, "weind"].sum()
    pct_miskin = n_miskin / n_total * 100 if n_total > 0 else 0
    p1v = weighted_mean(grp["p1kako"], grp["weind"])
    p2v = weighted_mean(grp["p2kako"], grp["weind"])
    rows_t1.append(
        {
            "Wilayah (r102)": int(r102_val),
            "N Miskin": round(n_miskin, 0),
            "% Miskin (P0)": round(pct_miskin, 2),
            "P1": round(p1v, 3),
            "P2": round(p2v, 3),
        }
    )
df_t1 = pd.DataFrame(rows_t1).set_index("Wilayah (r102)")
write_table(ws, "Tabel 1. Persentase, Kedalaman, dan Keparahan Kemiskinan", df_t1)

# ============================================================
# TABLE 2: Gender (r405) by Poverty Status
# ============================================================

ws = get_ws("T2_Gender")
sub = df_ind.copy()
result_rows = []
for sex_val in sorted(sub["r405"].dropna().unique()):
    row = {"Jenis Kelamin": "Laki-laki" if sex_val == 1 else "Perempuan"}
    for mk_val in [0, 1]:
        grp = sub[(sub["r405"] == sex_val) & (sub["mkako"] == mk_val)]
        row[f"N mkako={int(mk_val)}"] = round(grp["fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Jenis Kelamin": "Total"}
for mk_val in [0, 1]:
    total_row[f"N mkako={int(mk_val)}"] = sum(r[f"N mkako={int(mk_val)}"] for r in result_rows)
result_rows.append(total_row)
df_t2 = pd.DataFrame(result_rows).set_index("Jenis Kelamin")
# Calculate percentages using helper function
row_order = ["Laki-laki", "Perempuan"]
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t2, n_cols, pct_cols, row_order)
# Reorder columns: Miskin (mkako=1) first
cols_order_mk = ["N mkako=1", "N mkako=0", "% mkako=1", "% mkako=0"]
df_t2 = df_t2[cols_order_mk]
write_table(ws, "Tabel 2. Penduduk Menurut Jenis Kelamin dan Status Kemiskinan", df_t2)

# ============================================================
# TABLE 3: Age Group by Poverty (Laki-laki)
# ============================================================

ws = get_ws("T3_AgeGroup_Male")
sub = df_ind[df_ind["r405"] == 1].copy()  # Laki-laki
result_rows = []
age_labels = ["0-14", "15-64", "65+"]
for age_grp in age_labels:
    row = {"Kelompok Umur": age_grp}
    for mk_val in [0, 1]:
        mask = (sub["kelum1"].astype(str) == age_grp) & (sub["mkako"] == mk_val)
        row[f"N mkako={int(mk_val)}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Kelompok Umur": "Total"}
for mk_val in [0, 1]:
    total_row[f"N mkako={int(mk_val)}"] = sum(r[f"N mkako={int(mk_val)}"] for r in result_rows)
result_rows.append(total_row)
df_t3 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
# Calculate percentages using helper function
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t3, n_cols, pct_cols, age_labels)
df_t3 = df_t3[cols_order_mk]
write_table(
    ws,
    "Tabel 3. Persentase Penduduk Laki-laki Menurut Kelompok Umur dan Status Kemiskinan",
    df_t3,
)

# ============================================================
# TABLE 4: Age Group by Poverty (Perempuan)
# ============================================================

ws = get_ws("T4_AgeGroup_Female")
sub = df_ind[df_ind["r405"] == 2].copy()  # Perempuan
result_rows = []
for age_grp in age_labels:
    row = {"Kelompok Umur": age_grp}
    for mk_val in [0, 1]:
        mask = (sub["kelum1"].astype(str) == age_grp) & (sub["mkako"] == mk_val)
        row[f"N mkako={int(mk_val)}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Kelompok Umur": "Total"}
for mk_val in [0, 1]:
    total_row[f"N mkako={int(mk_val)}"] = sum(r[f"N mkako={int(mk_val)}"] for r in result_rows)
result_rows.append(total_row)
df_t4 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
# Calculate percentages using helper function
calc_pct_ensure_100(df_t4, n_cols, pct_cols, age_labels)
df_t4 = df_t4[cols_order_mk]
write_table(
    ws,
    "Tabel 4. Persentase Penduduk Perempuan Menurut Kelompok Umur dan Status Kemiskinan",
    df_t4,
)

# ============================================================
# TABLE 5: Age Group by Poverty (Total)
# ============================================================

ws = get_ws("T5_AgeGroup_Total")
sub = df_ind.copy()  # All genders
result_rows = []
for age_grp in age_labels:
    row = {"Kelompok Umur": age_grp}
    for mk_val in [0, 1]:
        mask = (sub["kelum1"].astype(str) == age_grp) & (sub["mkako"] == mk_val)
        row[f"N mkako={int(mk_val)}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Kelompok Umur": "Total"}
for mk_val in [0, 1]:
    total_row[f"N mkako={int(mk_val)}"] = sum(r[f"N mkako={int(mk_val)}"] for r in result_rows)
result_rows.append(total_row)
df_t5 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
# Calculate percentages using helper function
calc_pct_ensure_100(df_t5, n_cols, pct_cols, age_labels)
df_t5 = df_t5[cols_order_mk]
write_table(
    ws,
    "Tabel 5. Persentase Penduduk Menurut Kelompok Umur dan Status Kemiskinan",
    df_t5,
)

# ============================================================
# TABLE 6: Marital Status x Age Group by Poverty (Laki-laki, age >= 10)
# ============================================================

ws = get_ws("T6_MaritalStatus_Male")
sub = df_ind[
    (df_ind["r405"] == 1) & (df_ind["r407"] >= 10)
].copy()  # Laki-laki, age >= 10

marital_labels = {1: "Belum Kawin", 2: "Kawin", 3: "Cerai Hidup", 4: "Cerai Mati"}
age_groups = ["10-18", ">18"]

result_rows = []
for marital_val in [1, 2, 3, 4]:
    marital_label = marital_labels[marital_val]
    for age_grp in age_groups:
        row = {"Status Perkawinan": marital_label, "Kelompok Umur": age_grp}
        for mk in [0, 1]:
            mask = (
                (sub["r404"] == marital_val)
                & (sub["kelum2"].astype(str) == age_grp)
                & (sub["mkako"] == mk)
            )
            row[f"% mkako={mk}"] = 0  # Will calculate later
            row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
        result_rows.append(row)

# Calculate percentages within each marital status for each poverty column
for marital_val in [1, 2, 3, 4]:
    marital_label = marital_labels[marital_val]
    for mk in [0, 1]:
        # Sum of this marital status for this poverty status
        total_marital_mk = sum(
            r[f"N mkako={mk}"]
            for r in result_rows
            if r["Status Perkawinan"] == marital_label
        )
        # Calculate percentages
        for r in result_rows:
            if r["Status Perkawinan"] == marital_label and total_marital_mk > 0:
                r[f"% mkako={mk}"] = round(
                    r[f"N mkako={mk}"] / total_marital_mk * 100, 2
                )

# Add single Total row showing 100 for each column
total_row = {"Status Perkawinan": "Total", "Kelompok Umur": ""}
for mk in [0, 1]:
    total_row[f"% mkako={mk}"] = 100.00
    total_row[f"N mkako={mk}"] = ""
result_rows.append(total_row)

df_t6 = pd.DataFrame(result_rows)
# Reorder columns
cols_order = ["Status Perkawinan", "Kelompok Umur", "% mkako=1", "% mkako=0"]
df_t6 = df_t6[cols_order]
write_table(
    ws,
    "Tabel 6. Persentase Penduduk Laki-laki Berumur 10+ Menurut Status Perkawinan, Kelompok Umur dan Status Kemiskinan",
    df_t6,
)

# ============================================================
# TABLE 7: Marital Status x Age Group by Poverty (Perempuan, age >= 10)
# ============================================================

ws = get_ws("T7_MaritalStatus_Female")
sub = df_ind[
    (df_ind["r405"] == 2) & (df_ind["r407"] >= 10)
].copy()  # Perempuan, age >= 10

result_rows = []
for marital_val in [1, 2, 3, 4]:
    marital_label = marital_labels[marital_val]
    for age_grp in age_groups:
        row = {"Status Perkawinan": marital_label, "Kelompok Umur": age_grp}
        for mk in [0, 1]:
            mask = (
                (sub["r404"] == marital_val)
                & (sub["kelum2"].astype(str) == age_grp)
                & (sub["mkako"] == mk)
            )
            row[f"% mkako={mk}"] = 0
            row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
        result_rows.append(row)

# Calculate percentages within each marital status
for marital_val in [1, 2, 3, 4]:
    marital_label = marital_labels[marital_val]
    for mk in [0, 1]:
        total_marital_mk = sum(
            r[f"N mkako={mk}"]
            for r in result_rows
            if r["Status Perkawinan"] == marital_label
        )
        for r in result_rows:
            if r["Status Perkawinan"] == marital_label and total_marital_mk > 0:
                r[f"% mkako={mk}"] = round(
                    r[f"N mkako={mk}"] / total_marital_mk * 100, 2
                )

# Add single Total row showing 100 for each column
total_row = {"Status Perkawinan": "Total", "Kelompok Umur": ""}
for mk in [0, 1]:
    total_row[f"% mkako={mk}"] = 100.00
    total_row[f"N mkako={mk}"] = ""
result_rows.append(total_row)

df_t7 = pd.DataFrame(result_rows)
df_t7 = df_t7[cols_order]
write_table(
    ws,
    "Tabel 7. Persentase Penduduk Perempuan Berumur 10+ Menurut Status Perkawinan, Kelompok Umur dan Status Kemiskinan",
    df_t7,
)

# ============================================================
# TABLE 8: Marital Status x Age Group by Poverty (Total, age >= 10)
# ============================================================

ws = get_ws("T8_MaritalStatus_Total")
sub = df_ind[df_ind["r407"] >= 10].copy()  # All genders, age >= 10

result_rows = []
for marital_val in [1, 2, 3, 4]:
    marital_label = marital_labels[marital_val]
    for age_grp in age_groups:
        row = {"Status Perkawinan": marital_label, "Kelompok Umur": age_grp}
        for mk in [0, 1]:
            mask = (
                (sub["r404"] == marital_val)
                & (sub["kelum2"].astype(str) == age_grp)
                & (sub["mkako"] == mk)
            )
            row[f"% mkako={mk}"] = 0
            row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
        result_rows.append(row)

# Calculate percentages within each marital status
for marital_val in [1, 2, 3, 4]:
    marital_label = marital_labels[marital_val]
    for mk in [0, 1]:
        total_marital_mk = sum(
            r[f"N mkako={mk}"]
            for r in result_rows
            if r["Status Perkawinan"] == marital_label
        )
        for r in result_rows:
            if r["Status Perkawinan"] == marital_label and total_marital_mk > 0:
                r[f"% mkako={mk}"] = round(
                    r[f"N mkako={mk}"] / total_marital_mk * 100, 2
                )

# Add single Total row showing 100 for each column
total_row = {"Status Perkawinan": "Total", "Kelompok Umur": ""}
for mk in [0, 1]:
    total_row[f"% mkako={mk}"] = 100.00
    total_row[f"N mkako={mk}"] = ""
result_rows.append(total_row)

df_t8 = pd.DataFrame(result_rows)
df_t8 = df_t8[cols_order]
write_table(
    ws,
    "Tabel 8. Persentase Penduduk Berumur 10+ Menurut Status Perkawinan, Kelompok Umur dan Status Kemiskinan",
    df_t8,
)

# ============================================================
# TABLE 9: Household Head Gender by Poverty
# ============================================================

ws = get_ws("T9_HHHead_Gender")
sub = df_ind[df_ind["r403"] == 1].copy()
result_rows = []
for sex in sorted(sub["r405"].dropna().unique()):
    row = {"Jenis Kelamin": "Laki-laki" if sex == 1 else "Perempuan"}
    for mk in [0, 1]:
        mask = (sub["r405"] == sex) & (sub["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Jenis Kelamin": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={int(mk)}"] = sum(r[f"N mkako={int(mk)}"] for r in result_rows)
result_rows.append(total_row)
df_t9 = pd.DataFrame(result_rows).set_index("Jenis Kelamin")
# Calculate percentages (column-wise) and ensure they sum to 100
for mk in [0, 1]:
    n_col = f"N mkako={mk}"
    pct_col = f"% mkako={mk}"
    total_n = df_t9.loc["Total", n_col]
    if total_n > 0:
        # Calculate percentages for non-Total rows
        pct_sum = 0
        non_total_indices = [idx for idx in df_t9.index if idx != "Total"]
        for idx in non_total_indices[:-1]:  # All except last and Total
            pct = round(df_t9.loc[idx, n_col] / total_n * 100, 2)
            df_t9.loc[idx, pct_col] = pct
            pct_sum += pct
        # Last non-Total row: make it sum to 100
        last_idx = non_total_indices[-1]
        df_t9.loc[last_idx, pct_col] = round(100 - pct_sum, 2)
        # Total row
        df_t9.loc["Total", pct_col] = 100.00
# Reorder columns: Miskin (mkako=1) first, then Tidak Miskin (mkako=0)
cols_order = ["N mkako=1", "N mkako=0", "% mkako=1", "% mkako=0"]
df_t9 = df_t9[cols_order]
write_table(
    ws,
    "Tabel 9. Persentase Kepala Rumah Tangga Menurut Status Miskin dan Jenis Kelamin",
    df_t9,
)

# ============================================================
# TABLE 10: HH Head Marital Status by Poverty (Laki-laki)
# ============================================================

ws = get_ws("T10_HHHead_Marital_Male")
sub = df_ind[(df_ind["r403"] == 1) & (df_ind["r405"] == 1)].copy()  # Laki-laki

marital_labels = {1: "Belum Kawin", 2: "Kawin", 3: "Cerai Hidup", 4: "Cerai Mati"}
marital_order = [1, 2, 3, 4]

result_rows = []
for marital_val in marital_order:
    row = {"Status Perkawinan": marital_labels[marital_val]}
    for mk in [0, 1]:
        mask = (sub["r404"] == marital_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Status Perkawinan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t10 = pd.DataFrame(result_rows).set_index("Status Perkawinan")
# Calculate percentages (column-wise) ensuring sum = 100
for mk in [0, 1]:
    n_col = f"N mkako={mk}"
    total_n = df_t10.loc["Total", n_col]
    pct_col = f"% mkako={mk}"
    if total_n > 0:
        # Calculate for all except last and Total
        pct_sum = 0
        for idx in marital_order[:-1]:  # All except last
            label = marital_labels[idx]
            pct = round(df_t10.loc[label, n_col] / total_n * 100, 2)
            df_t10.loc[label, pct_col] = pct
            pct_sum += pct
        # Last item: make it sum to 100
        last_label = marital_labels[marital_order[-1]]
        df_t10.loc[last_label, pct_col] = round(100 - pct_sum, 2)
        # Total row
        df_t10.loc["Total", pct_col] = 100.00
# Reorder columns: Miskin first
cols_order = ["N mkako=1", "N mkako=0", "% mkako=1", "% mkako=0"]
df_t10 = df_t10[cols_order]
write_table(
    ws,
    "Tabel 10. Persentase Kepala Rumah Tangga Laki-laki Menurut Status Miskin dan Status Perkawinan",
    df_t10,
)

# ============================================================
# TABLE 11: HH Head Marital Status by Poverty (Perempuan)
# ============================================================

ws = get_ws("T11_HHHead_Marital_Female")
sub = df_ind[(df_ind["r403"] == 1) & (df_ind["r405"] == 2)].copy()  # Perempuan

result_rows = []
for marital_val in marital_order:
    row = {"Status Perkawinan": marital_labels[marital_val]}
    for mk in [0, 1]:
        mask = (sub["r404"] == marital_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Status Perkawinan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t11 = pd.DataFrame(result_rows).set_index("Status Perkawinan")
# Calculate percentages (column-wise) ensuring sum = 100
for mk in [0, 1]:
    n_col = f"N mkako={mk}"
    total_n = df_t11.loc["Total", n_col]
    pct_col = f"% mkako={mk}"
    if total_n > 0:
        pct_sum = 0
        for idx in marital_order[:-1]:
            label = marital_labels[idx]
            pct = round(df_t11.loc[label, n_col] / total_n * 100, 2)
            df_t11.loc[label, pct_col] = pct
            pct_sum += pct
        last_label = marital_labels[marital_order[-1]]
        df_t11.loc[last_label, pct_col] = round(100 - pct_sum, 2)
        df_t11.loc["Total", pct_col] = 100.00
df_t11 = df_t11[cols_order]
write_table(
    ws,
    "Tabel 11. Persentase Kepala Rumah Tangga Perempuan Menurut Status Miskin dan Status Perkawinan",
    df_t11,
)

# ============================================================
# TABLE 12: HH Head Marital Status by Poverty (Total)
# ============================================================

ws = get_ws("T12_HHHead_Marital_Total")
sub = df_ind[df_ind["r403"] == 1].copy()  # All household heads

result_rows = []
for marital_val in marital_order:
    row = {"Status Perkawinan": marital_labels[marital_val]}
    for mk in [0, 1]:
        mask = (sub["r404"] == marital_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Status Perkawinan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t12 = pd.DataFrame(result_rows).set_index("Status Perkawinan")
# Calculate percentages (column-wise) ensuring sum = 100
for mk in [0, 1]:
    n_col = f"N mkako={mk}"
    total_n = df_t12.loc["Total", n_col]
    pct_col = f"% mkako={mk}"
    if total_n > 0:
        pct_sum = 0
        for idx in marital_order[:-1]:
            label = marital_labels[idx]
            pct = round(df_t12.loc[label, n_col] / total_n * 100, 2)
            df_t12.loc[label, pct_col] = pct
            pct_sum += pct
        last_label = marital_labels[marital_order[-1]]
        df_t12.loc[last_label, pct_col] = round(100 - pct_sum, 2)
        df_t12.loc["Total", pct_col] = 100.00
df_t12 = df_t12[cols_order]
write_table(
    ws,
    "Tabel 12. Persentase Kepala Rumah Tangga Menurut Status Miskin dan Status Perkawinan",
    df_t12,
)

# ============================================================
# TABLE 13: HH Head Age Group by Poverty (Laki-laki, age >= 15)
# ============================================================

ws = get_ws("T13_HHHead_Age_Male")
sub = df_ind[
    (df_ind["r403"] == 1) & (df_ind["r405"] == 1) & (df_ind["r407"] >= 15)
].copy()
age_labels = ["15-24", "25-44", "45-64", "65+"]

result_rows = []
for age_grp in age_labels:
    row = {"Kelompok Umur": age_grp}
    for mk in [0, 1]:
        mask = (sub["kelum3"].astype(str) == age_grp) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Kelompok Umur": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t13 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
# Calculate percentages (column-wise) ensuring sum = 100
for mk in [0, 1]:
    n_col = f"N mkako={mk}"
    total_n = df_t13.loc["Total", n_col]
    pct_col = f"% mkako={mk}"
    if total_n > 0:
        pct_sum = 0
        for idx in age_labels[:-1]:  # All except last
            pct = round(df_t13.loc[idx, n_col] / total_n * 100, 2)
            df_t13.loc[idx, pct_col] = pct
            pct_sum += pct
        # Last item: make it sum to 100
        df_t13.loc[age_labels[-1], pct_col] = round(100 - pct_sum, 2)
        df_t13.loc["Total", pct_col] = 100.00
# Reorder columns: Miskin first
df_t13 = df_t13[cols_order_mk]
write_table(
    ws,
    "Tabel 13. Persentase Kepala Rumah Tangga Laki-laki Berumur 15+ Menurut Status Miskin dan Kelompok Umur",
    df_t13,
)

# ============================================================
# TABLE 14: HH Head Age Group by Poverty (Perempuan, age >= 15)
# ============================================================

ws = get_ws("T14_HHHead_Age_Female")
sub = df_ind[
    (df_ind["r403"] == 1) & (df_ind["r405"] == 2) & (df_ind["r407"] >= 15)
].copy()

result_rows = []
for age_grp in age_labels:
    row = {"Kelompok Umur": age_grp}
    for mk in [0, 1]:
        mask = (sub["kelum3"].astype(str) == age_grp) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Kelompok Umur": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t14 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
# Calculate percentages (column-wise) ensuring sum = 100
for mk in [0, 1]:
    n_col = f"N mkako={mk}"
    total_n = df_t14.loc["Total", n_col]
    pct_col = f"% mkako={mk}"
    if total_n > 0:
        pct_sum = 0
        for idx in age_labels[:-1]:
            pct = round(df_t14.loc[idx, n_col] / total_n * 100, 2)
            df_t14.loc[idx, pct_col] = pct
            pct_sum += pct
        df_t14.loc[age_labels[-1], pct_col] = round(100 - pct_sum, 2)
        df_t14.loc["Total", pct_col] = 100.00
df_t14 = df_t14[cols_order_mk]
write_table(
    ws,
    "Tabel 14. Persentase Kepala Rumah Tangga Perempuan Berumur 15+ Menurut Status Miskin dan Kelompok Umur",
    df_t14,
)

# ============================================================
# TABLE 15: HH Head Age Group by Poverty (Total, age >= 15)
# ============================================================

ws = get_ws("T15_HHHead_Age_Total")
sub = df_ind[(df_ind["r403"] == 1) & (df_ind["r407"] >= 15)].copy()

result_rows = []
for age_grp in age_labels:
    row = {"Kelompok Umur": age_grp}
    for mk in [0, 1]:
        mask = (sub["kelum3"].astype(str) == age_grp) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Kelompok Umur": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t15 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
# Calculate percentages (column-wise) ensuring sum = 100
for mk in [0, 1]:
    n_col = f"N mkako={mk}"
    total_n = df_t15.loc["Total", n_col]
    pct_col = f"% mkako={mk}"
    if total_n > 0:
        pct_sum = 0
        for idx in age_labels[:-1]:
            pct = round(df_t15.loc[idx, n_col] / total_n * 100, 2)
            df_t15.loc[idx, pct_col] = pct
            pct_sum += pct
        df_t15.loc[age_labels[-1], pct_col] = round(100 - pct_sum, 2)
        df_t15.loc["Total", pct_col] = 100.00
df_t15 = df_t15[cols_order_mk]
write_table(
    ws,
    "Tabel 15. Persentase Kepala Rumah Tangga Berumur 15+ Menurut Status Miskin dan Kelompok Umur",
    df_t15,
)

# ============================================================
# TABLE 16: Household Size by Poverty
# ============================================================

ws = get_ws("T16_HHSize")
sub = df_ind[df_ind["r403"] == 1].copy()
jart_labels = ["1-3", "4-6", ">=7"]
result_rows = []
for jart in jart_labels:
    row = {"Jumlah ART": jart}
    for mk in [0, 1]:
        mask = (sub["keljart"] == jart) & (sub["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row (sum of rounded values for consistency)
total_row = {"Jumlah ART": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t16 = pd.DataFrame(result_rows).set_index("Jumlah ART")
# Calculate percentages using helper function
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t16, n_cols, pct_cols, jart_labels)
df_t16 = df_t16[cols_order_mk]
write_table(
    ws,
    "Tabel 16. Persentase Rata-rata Banyaknya Anggota Rumah Tangga Menurut Status Miskin",
    df_t16,
)

# ============================================================
# TABLE 17: School Participation (r611) by Poverty (Laki-laki, age 5-24)
# ============================================================

ws = get_ws("T17_SchoolPart_Male")
sub = df_ind[
    (df_ind["r405"] == 1) & (df_ind["r407"] >= 5) & (df_ind["r407"] <= 24)
].copy()
r611_labels = {
    1: "Tidak/belum pernah bersekolah",
    2: "Masih bersekolah",
    3: "Tidak bersekolah lagi",
}
r611_order = [1, 2, 3]

result_rows = []
for r611_val in r611_order:
    row = {"Status Pendidikan": r611_labels[r611_val]}
    for mk in [0, 1]:
        mask = (sub["r611"] == r611_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Status Pendidikan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t17 = pd.DataFrame(result_rows).set_index("Status Pendidikan")
# Calculate percentages using helper function
row_order = [r611_labels[v] for v in r611_order]
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t17, n_cols, pct_cols, row_order)
df_t17 = df_t17[cols_order_mk]
write_table(
    ws,
    "Tabel 17. Persentase Penduduk Laki-laki Berumur 5-24 Tahun Menurut Status Miskin dan Status Pendidikan",
    df_t17,
)

# ============================================================
# TABLE 18: School Participation (r611) by Poverty (Perempuan, age 5-24)
# ============================================================

ws = get_ws("T18_SchoolPart_Female")
sub = df_ind[
    (df_ind["r405"] == 2) & (df_ind["r407"] >= 5) & (df_ind["r407"] <= 24)
].copy()

result_rows = []
for r611_val in r611_order:
    row = {"Status Pendidikan": r611_labels[r611_val]}
    for mk in [0, 1]:
        mask = (sub["r611"] == r611_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Status Pendidikan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t18 = pd.DataFrame(result_rows).set_index("Status Pendidikan")
# Calculate percentages using helper function
calc_pct_ensure_100(df_t18, n_cols, pct_cols, row_order)
df_t18 = df_t18[cols_order_mk]
write_table(
    ws,
    "Tabel 18. Persentase Penduduk Perempuan Berumur 5-24 Tahun Menurut Status Miskin dan Status Pendidikan",
    df_t18,
)

# ============================================================
# TABLE 19: School Participation (r611) by Poverty (Total, age 5-24)
# ============================================================

ws = get_ws("T19_SchoolPart_Total")
sub = df_ind[(df_ind["r407"] >= 5) & (df_ind["r407"] <= 24)].copy()

result_rows = []
for r611_val in r611_order:
    row = {"Status Pendidikan": r611_labels[r611_val]}
    for mk in [0, 1]:
        mask = (sub["r611"] == r611_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Status Pendidikan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t19 = pd.DataFrame(result_rows).set_index("Status Pendidikan")
# Calculate percentages using helper function
calc_pct_ensure_100(df_t19, n_cols, pct_cols, row_order)
df_t19 = df_t19[cols_order_mk]
write_table(
    ws,
    "Tabel 19. Persentase Penduduk Berumur 5-24 Tahun Menurut Status Miskin dan Status Pendidikan",
    df_t19,
)

# ============================================================
# TABLE 20: Education Level by Poverty (Laki-laki, age 15+)
# ============================================================

ws = get_ws("T20_EducLevel_Male")
sub = df_ind[(df_ind["r405"] == 1) & (df_ind["r407"] >= 15)].copy()

result_rows = []
for educ in educ_order:
    row = {"Pendidikan Tertinggi Yang Ditamatkan": educ}
    for mk in [0, 1]:
        mask = (sub["kelijasah"] == educ) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Pendidikan Tertinggi Yang Ditamatkan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t20 = pd.DataFrame(result_rows).set_index("Pendidikan Tertinggi Yang Ditamatkan")
# Calculate percentages using helper function
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t20, n_cols, pct_cols, educ_order)
df_t20 = df_t20[cols_order_mk]
write_table(
    ws,
    "Tabel 20. Persentase Penduduk Laki-laki Berumur 15+ Tahun Menurut Status Miskin dan Pendidikan Tertinggi Yang Ditamatkan",
    df_t20,
)

# ============================================================
# TABLE 21: Education Level by Poverty (Perempuan, age 15+)
# ============================================================

ws = get_ws("T21_EducLevel_Female")
sub = df_ind[(df_ind["r405"] == 2) & (df_ind["r407"] >= 15)].copy()

result_rows = []
for educ in educ_order:
    row = {"Pendidikan Tertinggi Yang Ditamatkan": educ}
    for mk in [0, 1]:
        mask = (sub["kelijasah"] == educ) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Pendidikan Tertinggi Yang Ditamatkan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t21 = pd.DataFrame(result_rows).set_index("Pendidikan Tertinggi Yang Ditamatkan")
# Calculate percentages using helper function
calc_pct_ensure_100(df_t21, n_cols, pct_cols, educ_order)
df_t21 = df_t21[cols_order_mk]
write_table(
    ws,
    "Tabel 21. Persentase Penduduk Perempuan Berumur 15+ Tahun Menurut Status Miskin dan Pendidikan Tertinggi Yang Ditamatkan",
    df_t21,
)

# ============================================================
# TABLE 22: Education Level by Poverty (Total, age 15+)
# ============================================================

ws = get_ws("T22_EducLevel_Total")
sub = df_ind[df_ind["r407"] >= 15].copy()

result_rows = []
for educ in educ_order:
    row = {"Pendidikan Tertinggi Yang Ditamatkan": educ}
    for mk in [0, 1]:
        mask = (sub["kelijasah"] == educ) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Pendidikan Tertinggi Yang Ditamatkan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t22 = pd.DataFrame(result_rows).set_index("Pendidikan Tertinggi Yang Ditamatkan")
# Calculate percentages using helper function
calc_pct_ensure_100(df_t22, n_cols, pct_cols, educ_order)
df_t22 = df_t22[cols_order_mk]
write_table(
    ws,
    "Tabel 22. Persentase Penduduk Berumur 15+ Tahun Menurut Status Miskin dan Pendidikan Tertinggi Yang Ditamatkan",
    df_t22,
)

# ============================================================
# TABLE 23: School Participation Rate (APS) by Age Group and Poverty (Laki-laki, age 7-18)
# ============================================================

ws = get_ws("T23_APS_Male")
sub = df_ind[
    (df_ind["r405"] == 1) & (df_ind["r407"] >= 7) & (df_ind["r407"] <= 18)
].copy()
age_groups_aps = ["7-12", "13-15", "16-18"]

result_rows = []
for age_grp in age_groups_aps:
    row = {"Kelompok Umur": age_grp}
    for mk in [0, 1]:
        mask = (sub["kelum4"].astype(str) == age_grp) & (sub["mkako"] == mk)
        grp = sub[mask]
        if len(grp) > 0:
            row[f"APS mkako={mk}"] = round(weighted_mean(grp["aps"], grp["fwt"]), 2)
        else:
            row[f"APS mkako={mk}"] = np.nan
    result_rows.append(row)

df_t23 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
# Reorder columns: Miskin first
cols_order_aps = ["APS mkako=1", "APS mkako=0"]
df_t23 = df_t23[cols_order_aps]
write_table(
    ws,
    "Tabel 23. Angka Partisipasi Sekolah (APS) Penduduk Laki-laki Berumur 7-18 Tahun Menurut Status Miskin",
    df_t23,
)

# ============================================================
# TABLE 24: School Participation Rate (APS) by Age Group and Poverty (Perempuan, age 7-18)
# ============================================================

ws = get_ws("T24_APS_Female")
sub = df_ind[
    (df_ind["r405"] == 2) & (df_ind["r407"] >= 7) & (df_ind["r407"] <= 18)
].copy()

result_rows = []
for age_grp in age_groups_aps:
    row = {"Kelompok Umur": age_grp}
    for mk in [0, 1]:
        mask = (sub["kelum4"].astype(str) == age_grp) & (sub["mkako"] == mk)
        grp = sub[mask]
        if len(grp) > 0:
            row[f"APS mkako={mk}"] = round(weighted_mean(grp["aps"], grp["fwt"]), 2)
        else:
            row[f"APS mkako={mk}"] = np.nan
    result_rows.append(row)

df_t24 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
df_t24 = df_t24[cols_order_aps]
write_table(
    ws,
    "Tabel 24. Angka Partisipasi Sekolah (APS) Penduduk Perempuan Berumur 7-18 Tahun Menurut Status Miskin",
    df_t24,
)

# ============================================================
# TABLE 25: School Participation Rate (APS) by Age Group and Poverty (Total, age 7-18)
# ============================================================

ws = get_ws("T25_APS_Total")
sub = df_ind[(df_ind["r407"] >= 7) & (df_ind["r407"] <= 18)].copy()

result_rows = []
for age_grp in age_groups_aps:
    row = {"Kelompok Umur": age_grp}
    for mk in [0, 1]:
        mask = (sub["kelum4"].astype(str) == age_grp) & (sub["mkako"] == mk)
        grp = sub[mask]
        if len(grp) > 0:
            row[f"APS mkako={mk}"] = round(weighted_mean(grp["aps"], grp["fwt"]), 2)
        else:
            row[f"APS mkako={mk}"] = np.nan
    result_rows.append(row)

df_t25 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
df_t25 = df_t25[cols_order_aps]
write_table(
    ws,
    "Tabel 25. Angka Partisipasi Sekolah (APS) Penduduk Berumur 7-18 Tahun Menurut Status Miskin",
    df_t25,
)

# ============================================================
# TABLE 26: Literacy Rate (AMH) by Age Group and Poverty (Laki-laki)
# ============================================================

ws = get_ws("T26_AMH_Male")
age_ranges_amh = [("15-24 tahun", 15, 24), ("15-55 tahun", 15, 55)]

result_rows = []
for age_label, lo, hi in age_ranges_amh:
    sub = df_ind[
        (df_ind["r405"] == 1) & (df_ind["r407"] >= lo) & (df_ind["r407"] <= hi)
    ].copy()
    row = {"Kelompok Umur": age_label}
    for mk in [0, 1]:
        mask = sub["mkako"] == mk
        grp = sub[mask]
        if len(grp) > 0:
            row[f"AMH mkako={mk}"] = round(weighted_mean(grp["amh"], grp["fwt"]), 2)
        else:
            row[f"AMH mkako={mk}"] = np.nan
    result_rows.append(row)

df_t26 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
cols_order_amh = ["AMH mkako=1", "AMH mkako=0"]
df_t26 = df_t26[cols_order_amh]
write_table(
    ws,
    "Tabel 26. Angka Melek Huruf (AMH) Penduduk Laki-laki Menurut Status Miskin",
    df_t26,
)

# ============================================================
# TABLE 27: Literacy Rate (AMH) by Age Group and Poverty (Perempuan)
# ============================================================

ws = get_ws("T27_AMH_Female")

result_rows = []
for age_label, lo, hi in age_ranges_amh:
    sub = df_ind[
        (df_ind["r405"] == 2) & (df_ind["r407"] >= lo) & (df_ind["r407"] <= hi)
    ].copy()
    row = {"Kelompok Umur": age_label}
    for mk in [0, 1]:
        mask = sub["mkako"] == mk
        grp = sub[mask]
        if len(grp) > 0:
            row[f"AMH mkako={mk}"] = round(weighted_mean(grp["amh"], grp["fwt"]), 2)
        else:
            row[f"AMH mkako={mk}"] = np.nan
    result_rows.append(row)

df_t27 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
df_t27 = df_t27[cols_order_amh]
write_table(
    ws,
    "Tabel 27. Angka Melek Huruf (AMH) Penduduk Perempuan Menurut Status Miskin",
    df_t27,
)

# ============================================================
# TABLE 28: Literacy Rate (AMH) by Age Group and Poverty (Total)
# ============================================================

ws = get_ws("T28_AMH_Total")

result_rows = []
for age_label, lo, hi in age_ranges_amh:
    sub = df_ind[(df_ind["r407"] >= lo) & (df_ind["r407"] <= hi)].copy()
    row = {"Kelompok Umur": age_label}
    for mk in [0, 1]:
        mask = sub["mkako"] == mk
        grp = sub[mask]
        if len(grp) > 0:
            row[f"AMH mkako={mk}"] = round(weighted_mean(grp["amh"], grp["fwt"]), 2)
        else:
            row[f"AMH mkako={mk}"] = np.nan
    result_rows.append(row)

df_t28 = pd.DataFrame(result_rows).set_index("Kelompok Umur")
df_t28 = df_t28[cols_order_amh]
write_table(
    ws,
    "Tabel 28. Angka Melek Huruf (AMH) Penduduk Menurut Status Miskin",
    df_t28,
)

# ============================================================
# TABLE 29: HH Head Education by Poverty (Laki-laki)
# ============================================================

ws = get_ws("T29_HHHead_Educ_Male")
sub = df_ind[(df_ind["r403"] == 1) & (df_ind["r405"] == 1)].copy()  # Male HH heads

result_rows = []
for educ in educ_order:
    row = {"Pendidikan Tertinggi Yang Ditamatkan": educ}
    for mk in [0, 1]:
        mask = (sub["kelijasah"] == educ) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Pendidikan Tertinggi Yang Ditamatkan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t29 = pd.DataFrame(result_rows).set_index("Pendidikan Tertinggi Yang Ditamatkan")
# Calculate percentages (column-wise) ensuring sum = 100
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t29, n_cols, pct_cols, educ_order)
df_t29 = df_t29[cols_order_mk]
write_table(
    ws,
    "Tabel 29. Persentase Kepala Rumah Tangga Laki-laki Menurut Status Miskin dan Pendidikan Tertinggi Yang Ditamatkan",
    df_t29,
)

# ============================================================
# TABLE 30: HH Head Education by Poverty (Perempuan)
# ============================================================

ws = get_ws("T30_HHHead_Educ_Female")
sub = df_ind[(df_ind["r403"] == 1) & (df_ind["r405"] == 2)].copy()  # Female HH heads

result_rows = []
for educ in educ_order:
    row = {"Pendidikan Tertinggi Yang Ditamatkan": educ}
    for mk in [0, 1]:
        mask = (sub["kelijasah"] == educ) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Pendidikan Tertinggi Yang Ditamatkan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t30 = pd.DataFrame(result_rows).set_index("Pendidikan Tertinggi Yang Ditamatkan")
# Calculate percentages (column-wise) ensuring sum = 100
calc_pct_ensure_100(df_t30, n_cols, pct_cols, educ_order)
df_t30 = df_t30[cols_order_mk]
write_table(
    ws,
    "Tabel 30. Persentase Kepala Rumah Tangga Perempuan Menurut Status Miskin dan Pendidikan Tertinggi Yang Ditamatkan",
    df_t30,
)

# ============================================================
# TABLE 31: HH Head Education by Poverty (Total)
# ============================================================

ws = get_ws("T31_HHHead_Educ_Total")
sub = df_ind[df_ind["r403"] == 1].copy()  # All HH heads

result_rows = []
for educ in educ_order:
    row = {"Pendidikan Tertinggi Yang Ditamatkan": educ}
    for mk in [0, 1]:
        mask = (sub["kelijasah"] == educ) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
# Add Total row
total_row = {"Pendidikan Tertinggi Yang Ditamatkan": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t31 = pd.DataFrame(result_rows).set_index("Pendidikan Tertinggi Yang Ditamatkan")
# Calculate percentages (column-wise) ensuring sum = 100
calc_pct_ensure_100(df_t31, n_cols, pct_cols, educ_order)
df_t31 = df_t31[cols_order_mk]
write_table(
    ws,
    "Tabel 31. Persentase Kepala Rumah Tangga Menurut Status Miskin, Jenis Kelamin Dan Pendidikan Tertinggi Yang Ditamatkan",
    df_t31,
)

# ============================================================
# TABLE 32: Employment Status by Poverty (Laki-laki, 15+)
# ============================================================

ws = get_ws("T32_Employment_Male")
sub = df_ind[(df_ind["r405"] == 1) & (df_ind["r407"] >= 15)].copy()

kerja_labels = [
    "Tidak Bekerja",
    "Bekerja di Sektor Formal",
    "Bekerja di Sektor Informal",
]

result_rows = []
for kerja in kerja_labels:
    row = {"Status Bekerja": kerja}
    for mk in [0, 1]:
        if kerja == "Tidak Bekerja":
            mask = (sub["tkerja"] == 100) & (sub["mkako"] == mk)
        elif kerja == "Bekerja di Sektor Formal":
            mask = (
                (sub["statuskerja"] == "Formal")
                & (sub["tkerja"] == 0)
                & (sub["mkako"] == mk)
            )
        else:
            mask = (
                (sub["statuskerja"] == "Informal")
                & (sub["tkerja"] == 0)
                & (sub["mkako"] == mk)
            )
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Status Bekerja": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t32 = pd.DataFrame(result_rows).set_index("Status Bekerja")
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t32, n_cols, pct_cols, kerja_labels)
df_t32 = df_t32[cols_order_mk]
write_table(
    ws,
    "Tabel 32. Persentase Penduduk Laki-laki Berumur 15 Tahun Keatas Menurut Status Bekerja dan Status Miskin",
    df_t32,
)

# ============================================================
# TABLE 33: Employment Status by Poverty (Perempuan, 15+)
# ============================================================

ws = get_ws("T33_Employment_Female")
sub = df_ind[(df_ind["r405"] == 2) & (df_ind["r407"] >= 15)].copy()

result_rows = []
for kerja in kerja_labels:
    row = {"Status Bekerja": kerja}
    for mk in [0, 1]:
        if kerja == "Tidak Bekerja":
            mask = (sub["tkerja"] == 100) & (sub["mkako"] == mk)
        elif kerja == "Bekerja di Sektor Formal":
            mask = (
                (sub["statuskerja"] == "Formal")
                & (sub["tkerja"] == 0)
                & (sub["mkako"] == mk)
            )
        else:
            mask = (
                (sub["statuskerja"] == "Informal")
                & (sub["tkerja"] == 0)
                & (sub["mkako"] == mk)
            )
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Status Bekerja": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t33 = pd.DataFrame(result_rows).set_index("Status Bekerja")
calc_pct_ensure_100(df_t33, n_cols, pct_cols, kerja_labels)
df_t33 = df_t33[cols_order_mk]
write_table(
    ws,
    "Tabel 33. Persentase Penduduk Perempuan Berumur 15 Tahun Keatas Menurut Status Bekerja dan Status Miskin",
    df_t33,
)

# ============================================================
# TABLE 34: Employment Status by Poverty (Total, 15+)
# ============================================================

ws = get_ws("T34_Employment_Total")
sub = df_ind[df_ind["r407"] >= 15].copy()

result_rows = []
for kerja in kerja_labels:
    row = {"Status Bekerja": kerja}
    for mk in [0, 1]:
        if kerja == "Tidak Bekerja":
            mask = (sub["tkerja"] == 100) & (sub["mkako"] == mk)
        elif kerja == "Bekerja di Sektor Formal":
            mask = (
                (sub["statuskerja"] == "Formal")
                & (sub["tkerja"] == 0)
                & (sub["mkako"] == mk)
            )
        else:
            mask = (
                (sub["statuskerja"] == "Informal")
                & (sub["tkerja"] == 0)
                & (sub["mkako"] == mk)
            )
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Status Bekerja": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t34 = pd.DataFrame(result_rows).set_index("Status Bekerja")
calc_pct_ensure_100(df_t34, n_cols, pct_cols, kerja_labels)
df_t34 = df_t34[cols_order_mk]
write_table(
    ws,
    "Tabel 34. Persentase Penduduk Berumur 15 Tahun Keatas Menurut Jenis Kelamin, Status Bekerja dan Status Miskin",
    df_t34,
)

# ============================================================
# TABLE 35: Employment Sector by Poverty (Laki-laki, 15+)
# ============================================================

ws = get_ws("T35_Sector_Male")
sub = df_ind[(df_ind["r405"] == 1) & (df_ind["r407"] >= 15)].copy()

sektor_labels = ["Tidak Bekerja", "Bekerja di Sektor Pertanian", "Bekerja Bukan di Sektor Pertanian"]

result_rows = []
for sektor in sektor_labels:
    row = {"Sektor Bekerja": sektor}
    for mk in [0, 1]:
        if sektor == "Tidak Bekerja":
            mask = (sub["tkerja"] == 100) & (sub["mkako"] == mk)
        elif sektor == "Bekerja di Sektor Pertanian":
            mask = (sub["sektorkerja"] == "Pertanian") & (sub["tkerja"] == 0) & (sub["mkako"] == mk)
        else:
            mask = (sub["sektorkerja"] == "Non Pertanian") & (sub["tkerja"] == 0) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Sektor Bekerja": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t35 = pd.DataFrame(result_rows).set_index("Sektor Bekerja")
calc_pct_ensure_100(df_t35, n_cols, pct_cols, sektor_labels)
df_t35 = df_t35[cols_order_mk]
write_table(
    ws,
    "Tabel 35. Persentase Penduduk Laki-laki Berumur 15 Tahun Keatas Menurut Sektor Bekerja dan Status Miskin",
    df_t35,
)

# ============================================================
# TABLE 36: Employment Sector by Poverty (Perempuan, 15+)
# ============================================================

ws = get_ws("T36_Sector_Female")
sub = df_ind[(df_ind["r405"] == 2) & (df_ind["r407"] >= 15)].copy()

result_rows = []
for sektor in sektor_labels:
    row = {"Sektor Bekerja": sektor}
    for mk in [0, 1]:
        if sektor == "Tidak Bekerja":
            mask = (sub["tkerja"] == 100) & (sub["mkako"] == mk)
        elif sektor == "Bekerja di Sektor Pertanian":
            mask = (sub["sektorkerja"] == "Pertanian") & (sub["tkerja"] == 0) & (sub["mkako"] == mk)
        else:
            mask = (sub["sektorkerja"] == "Non Pertanian") & (sub["tkerja"] == 0) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Sektor Bekerja": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t36 = pd.DataFrame(result_rows).set_index("Sektor Bekerja")
calc_pct_ensure_100(df_t36, n_cols, pct_cols, sektor_labels)
df_t36 = df_t36[cols_order_mk]
write_table(
    ws,
    "Tabel 36. Persentase Penduduk Perempuan Berumur 15 Tahun Keatas Menurut Sektor Bekerja dan Status Miskin",
    df_t36,
)

# ============================================================
# TABLE 37: Employment Sector by Poverty (Total, 15+)
# ============================================================

ws = get_ws("T37_Sector_Total")
sub = df_ind[df_ind["r407"] >= 15].copy()

result_rows = []
for sektor in sektor_labels:
    row = {"Sektor Bekerja": sektor}
    for mk in [0, 1]:
        if sektor == "Tidak Bekerja":
            mask = (sub["tkerja"] == 100) & (sub["mkako"] == mk)
        elif sektor == "Bekerja di Sektor Pertanian":
            mask = (sub["sektorkerja"] == "Pertanian") & (sub["tkerja"] == 0) & (sub["mkako"] == mk)
        else:
            mask = (sub["sektorkerja"] == "Non Pertanian") & (sub["tkerja"] == 0) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Sektor Bekerja": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t37 = pd.DataFrame(result_rows).set_index("Sektor Bekerja")
calc_pct_ensure_100(df_t37, n_cols, pct_cols, sektor_labels)
df_t37 = df_t37[cols_order_mk]
write_table(
    ws,
    "Tabel 37. Persentase Penduduk Berumur 15 Tahun Keatas Menurut Jenis Kelamin, Sektor Bekerja, dan Status Miskin",
    df_t37,
)

# ============================================================
# TABLE 38: JKN Ownership by Poverty (Laki-laki)
# ============================================================

ws = get_ws("T38_JKN_Male")
sub = df_ind[df_ind["r405"] == 1].copy()

jkn_labels = ["Ya", "Tidak"]

result_rows = []
for jkn in jkn_labels:
    row = {"Apakah Mempunyai Jaminan Kesehatan ?": jkn}
    for mk in [0, 1]:
        jkn_val = 1 if jkn == "Ya" else 0
        mask = (sub["milikjkn"] == jkn_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Apakah Mempunyai Jaminan Kesehatan ?": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t38 = pd.DataFrame(result_rows).set_index("Apakah Mempunyai Jaminan Kesehatan ?")
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t38, n_cols, pct_cols, jkn_labels)
df_t38 = df_t38[cols_order_mk]
write_table(
    ws,
    "Tabel 38. Persentase Penduduk Laki-laki yang Mempunyai Jaminan Kesehatan Menurut Status Miskin",
    df_t38,
)

# ============================================================
# TABLE 39: JKN Ownership by Poverty (Perempuan)
# ============================================================

ws = get_ws("T39_JKN_Female")
sub = df_ind[df_ind["r405"] == 2].copy()

result_rows = []
for jkn in jkn_labels:
    row = {"Apakah Mempunyai Jaminan Kesehatan ?": jkn}
    for mk in [0, 1]:
        jkn_val = 1 if jkn == "Ya" else 0
        mask = (sub["milikjkn"] == jkn_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Apakah Mempunyai Jaminan Kesehatan ?": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t39 = pd.DataFrame(result_rows).set_index("Apakah Mempunyai Jaminan Kesehatan ?")
calc_pct_ensure_100(df_t39, n_cols, pct_cols, jkn_labels)
df_t39 = df_t39[cols_order_mk]
write_table(
    ws,
    "Tabel 39. Persentase Penduduk Perempuan yang Mempunyai Jaminan Kesehatan Menurut Status Miskin",
    df_t39,
)

# ============================================================
# TABLE 40: JKN Ownership by Poverty (Total)
# ============================================================

ws = get_ws("T40_JKN_Total")
sub = df_ind.copy()

result_rows = []
for jkn in jkn_labels:
    row = {"Apakah Mempunyai Jaminan Kesehatan ?": jkn}
    for mk in [0, 1]:
        jkn_val = 1 if jkn == "Ya" else 0
        mask = (sub["milikjkn"] == jkn_val) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Apakah Mempunyai Jaminan Kesehatan ?": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t40 = pd.DataFrame(result_rows).set_index("Apakah Mempunyai Jaminan Kesehatan ?")
calc_pct_ensure_100(df_t40, n_cols, pct_cols, jkn_labels)
df_t40 = df_t40[cols_order_mk]
write_table(
    ws,
    "Tabel 40. Persentase Penduduk yang Mempunyai Jaminan Kesehatan Menurut Jenis Kelamin dan Status Miskin",
    df_t40,
)

# ============================================================
# TABLE 41: Smoking by Poverty (Laki-laki, 5+)
# ============================================================

ws = get_ws("T41_Smoking_Male")
sub = df_ind[(df_ind["r405"] == 1) & (df_ind["r407"] >= 5)].copy()

rokok_labels = ["Ya, setiap hari", "Ya, tidak setiap hari", "Tidak/Tidak Tahu"]

result_rows = []
for rk in rokok_labels:
    row = {"Apakah Selama Sebulan Terakhir Merokok Tembakau": rk}
    for mk in [0, 1]:
        mask = (sub["rokok"] == rk) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Apakah Selama Sebulan Terakhir Merokok Tembakau": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t41 = pd.DataFrame(result_rows).set_index("Apakah Selama Sebulan Terakhir Merokok Tembakau")
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t41, n_cols, pct_cols, rokok_labels)
df_t41 = df_t41[cols_order_mk]
write_table(
    ws,
    "Tabel 41. Persentase Penduduk Laki-laki Berumur 5 Tahun Keatas Menurut Apakah Selama Sebulan Terakhir Merokok Tembakau dan Status Miskin",
    df_t41,
)

# ============================================================
# TABLE 42: Smoking by Poverty (Perempuan, 5+)
# ============================================================

ws = get_ws("T42_Smoking_Female")
sub = df_ind[(df_ind["r405"] == 2) & (df_ind["r407"] >= 5)].copy()

result_rows = []
for rk in rokok_labels:
    row = {"Apakah Selama Sebulan Terakhir Merokok Tembakau": rk}
    for mk in [0, 1]:
        mask = (sub["rokok"] == rk) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Apakah Selama Sebulan Terakhir Merokok Tembakau": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t42 = pd.DataFrame(result_rows).set_index("Apakah Selama Sebulan Terakhir Merokok Tembakau")
calc_pct_ensure_100(df_t42, n_cols, pct_cols, rokok_labels)
df_t42 = df_t42[cols_order_mk]
write_table(
    ws,
    "Tabel 42. Persentase Penduduk Perempuan Berumur 5 Tahun Keatas Menurut Apakah Selama Sebulan Terakhir Merokok Tembakau dan Status Miskin",
    df_t42,
)

# ============================================================
# TABLE 43: Smoking by Poverty (Total, 5+)
# ============================================================

ws = get_ws("T43_Smoking_Total")
sub = df_ind[df_ind["r407"] >= 5].copy()

result_rows = []
for rk in rokok_labels:
    row = {"Apakah Selama Sebulan Terakhir Merokok Tembakau": rk}
    for mk in [0, 1]:
        mask = (sub["rokok"] == rk) & (sub["mkako"] == mk)
        row[f"N mkako={mk}"] = round(sub.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
total_row = {"Apakah Selama Sebulan Terakhir Merokok Tembakau": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)
df_t43 = pd.DataFrame(result_rows).set_index("Apakah Selama Sebulan Terakhir Merokok Tembakau")
calc_pct_ensure_100(df_t43, n_cols, pct_cols, rokok_labels)
df_t43 = df_t43[cols_order_mk]
write_table(
    ws,
    "Tabel 43. Persentase Penduduk Berumur 5 Tahun Keatas Menurut Jenis Kelamin, Apakah Selama Sebulan Terakhir Merokok Tembakau dan Status Miskin",
    df_t43,
)

# ============================================================
# TABLE 44: Water, Sanitation Access by Poverty (rt file)
# ============================================================

ws = get_ws("T44_Water_Sanitation")
result_rows = []

row = {"Penggunaan Air Minum dan Sanitasi": "Air Minum Layak"}
for mk in [0, 1]:
    grp = df_rt[df_rt["mkako"] == mk]
    w = grp["fwt"]
    row[f"% mkako={mk}"] = round(weighted_mean(grp["airmlayak"], w), 2)
result_rows.append(row)

row = {"Penggunaan Air Minum dan Sanitasi": "Air Minum Bersih"}
for mk in [0, 1]:
    grp = df_rt[df_rt["mkako"] == mk]
    w = grp["fwt"]
    row[f"% mkako={mk}"] = round(weighted_mean(grp["sab"], w), 2)
result_rows.append(row)

row = {"Penggunaan Air Minum dan Sanitasi": "Sanitasi Layak"}
for mk in [0, 1]:
    grp = df_rt[df_rt["mkako"] == mk]
    w = grp["fwt"]
    row[f"% mkako={mk}"] = round(weighted_mean(grp["sal"], w), 2)
result_rows.append(row)

df_t44 = pd.DataFrame(result_rows).set_index("Penggunaan Air Minum dan Sanitasi")
cols_order_mk_pct = ["% mkako=1", "% mkako=0"]
df_t44 = df_t44[cols_order_mk_pct]
write_table(
    ws, "Tabel 44. Persentase Rumah Tangga yang Menggunakan Air Minum Layak, Air Minum Bersih, Sanitasi Layak, dan Status Miskin", df_t44
)

# ============================================================
# TABLE 45: House Ownership by Poverty
# ============================================================

ws = get_ws("T45_HouseOwnership")

kepemilikan_labels = {1: "Milik Sendiri", 2: "Kontrak/Sewa", 3: "Bebas Sewa", 4: "Dinas"}
kepemilikan_order = [1, 2, 3, 4]

result_rows = []
for val in kepemilikan_order:
    row = {"Status Kepemilikan Rumah": kepemilikan_labels[val]}
    for mk in [0, 1]:
        mask = (df_rt["r1602"] == val) & (df_rt["mkako"] == mk)
        row[f"N mkako={mk}"] = round(df_rt.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)

total_row = {"Status Kepemilikan Rumah": "Total"}
for mk in [0, 1]:
    total_row[f"N mkako={mk}"] = sum(r[f"N mkako={mk}"] for r in result_rows)
result_rows.append(total_row)

df_t45 = pd.DataFrame(result_rows).set_index("Status Kepemilikan Rumah")
n_cols = ["N mkako=0", "N mkako=1"]
pct_cols = ["% mkako=0", "% mkako=1"]
calc_pct_ensure_100(df_t45, n_cols, pct_cols, [kepemilikan_labels[v] for v in kepemilikan_order])
df_t45 = df_t45[cols_order_mk]
write_table(ws, "Tabel 45. Persentase Rumah Tangga Menurut Status Kepemilikan Rumah dan Status Miskin", df_t45)

# ============================================================
# TABLE 46: Floor Area per Capita by Poverty
# ============================================================

ws = get_ws("T46_FloorArea")
result_rows = []
for klk in ["<=7,2 m2", ">7,2 m2"]:
    row = {"Luas Lantai per Kapita": klk}
    for mk in [0, 1]:
        mask = (df_rt["klkapita"].astype(str) == klk) & (df_rt["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(df_rt.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
df_t46 = pd.DataFrame(result_rows).set_index("Luas Lantai per Kapita")
for col in df_t46.columns:
    total = df_t46[col].sum()
    df_t46[col.replace("N ", "% ")] = (df_t46[col] / total * 100).round(2)
write_table(ws, "Tabel 46. Luas Lantai per Kapita Menurut Status Kemiskinan", df_t46)

# ============================================================
# TABLE 47: Roof Type by Poverty
# ============================================================

ws = get_ws("T47_RoofType")
roof_cats = ["Beton/Genteng", "Seng", "Asbes", "Bambu/kayu/jerami/lainnya"]
result_rows = []
for rk in roof_cats:
    row = {"Jenis Atap": rk}
    for mk in [0, 1]:
        mask = (df_rt["katap"] == rk) & (df_rt["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(df_rt.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
df_t47 = pd.DataFrame(result_rows).set_index("Jenis Atap")
for col in df_t47.columns:
    total = df_t47[col].sum()
    df_t47[col.replace("N ", "% ")] = (df_t47[col] / total * 100).round(2)
write_table(ws, "Tabel 47. Jenis Atap Rumah Terluas Menurut Status Kemiskinan", df_t47)

# ============================================================
# TABLE 48: Wall Type by Poverty
# ============================================================

ws = get_ws("T48_WallType")
wall_cats = ["Tembok", "Plesteran Anyaman Bambu/Kawat", "Kayu/Papan", "Lainnya"]
result_rows = []
for wk in wall_cats:
    row = {"Jenis Dinding": wk}
    for mk in [0, 1]:
        mask = (df_rt["kdinding"] == wk) & (df_rt["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(df_rt.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
df_t48 = pd.DataFrame(result_rows).set_index("Jenis Dinding")
for col in df_t48.columns:
    total = df_t48[col].sum()
    df_t48[col.replace("N ", "% ")] = (df_t48[col] / total * 100).round(2)
write_table(
    ws, "Tabel 48. Jenis Dinding Rumah Terluas Menurut Status Kemiskinan", df_t48
)

# ============================================================
# TABLE 49: Floor Type by Poverty
# ============================================================

ws = get_ws("T49_FloorType")
floor_cats = [
    "Marmer/Granit/Keramik/Parket/Vinil/Karpet",
    "Ubin/Tegel/Teraso",
    "Kayu/Papan",
    "Semen/Bata/Bambu/Tanah/Lainnya",
]
result_rows = []
for fk in floor_cats:
    row = {"Jenis Lantai": fk}
    for mk in [0, 1]:
        mask = (df_rt["klantai"] == fk) & (df_rt["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(df_rt.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
df_t49 = pd.DataFrame(result_rows).set_index("Jenis Lantai")
for col in df_t49.columns:
    total = df_t49[col].sum()
    df_t49[col.replace("N ", "% ")] = (df_t49[col] / total * 100).round(2)
write_table(
    ws, "Tabel 49. Jenis Lantai Rumah Terluas Menurut Status Kemiskinan", df_t49
)

# ============================================================
# TABLE 50: Electricity Source by Poverty
# ============================================================

ws = get_ws("T50_Electricity")
elec_cats = ["Listrik PLN", "Listrik Non PLN", "Bukan listrik"]
result_rows = []
for ek in elec_cats:
    row = {"Sumber Penerangan": ek}
    for mk in [0, 1]:
        mask = (df_rt["klistrik"] == ek) & (df_rt["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(df_rt.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
df_t50 = pd.DataFrame(result_rows).set_index("Sumber Penerangan")
for col in df_t50.columns:
    total = df_t50[col].sum()
    df_t50[col.replace("N ", "% ")] = (df_t50[col] / total * 100).round(2)
write_table(ws, "Tabel 50. Sumber Penerangan Utama Menurut Status Kemiskinan", df_t50)

# ============================================================
# TABLE 51-53: Food/Non-food Share by Gender and Poverty
# ============================================================

ws = get_ws("T51_53_FoodShare")
result_rows = []
for sex in sorted(df_ind["r405"].dropna().unique()):
    row = {"Jenis Kelamin": "Laki-laki" if sex == 1 else "Perempuan"}
    for mk in [0, 1]:
        mask = (df_ind["r405"] == sex) & (df_ind["mkako"] == mk)
        grp = df_ind[mask]
        w = grp["fwt"]
        row[f"Share Makanan mkako={int(mk)}"] = round(
            weighted_mean(grp["sharefoodkapita"], w), 2
        )
        row[f"Share Non-Makanan mkako={int(mk)}"] = round(
            weighted_mean(grp["sharenonfoodkapita"], w), 2
        )
    result_rows.append(row)
df_t5153 = pd.DataFrame(result_rows).set_index("Jenis Kelamin")
write_table(
    ws,
    "Tabel 51-53. Pangsa Pengeluaran Makanan/Non-Makanan Menurut Jenis Kelamin dan Status Kemiskinan",
    df_t5153,
)

# ============================================================
# TABLE 54: Food Per Capita by Poverty
# ============================================================

ws = get_ws("T54_FoodPerCapita")
result_rows = []
for mk in [0, 1]:
    grp = df_kp43[df_kp43["mkako"] == mk]
    result_rows.append(
        {
            "Status Kemiskinan": "Miskin" if mk == 1 else "Tidak Miskin",
            "Rata-rata Pengeluaran Makanan per Kapita (Rp)": round(
                weighted_mean(grp["foodkapita"], grp["weind"]), 0
            ),
        }
    )
# Total
result_rows.append(
    {
        "Status Kemiskinan": "Total",
        "Rata-rata Pengeluaran Makanan per Kapita (Rp)": round(
            weighted_mean(df_kp43["foodkapita"], df_kp43["weind"]), 0
        ),
    }
)
df_t54 = pd.DataFrame(result_rows).set_index("Status Kemiskinan")
write_table(
    ws,
    "Tabel 54. Rata-rata Pengeluaran Makanan per Kapita Menurut Status Kemiskinan",
    df_t54,
)

# ============================================================
# TABLE 55: Non-Food Per Capita by Poverty
# ============================================================

ws = get_ws("T55_NonFoodPerCapita")
result_rows = []
for mk in [0, 1]:
    grp = df_kp43[df_kp43["mkako"] == mk]
    result_rows.append(
        {
            "Status Kemiskinan": "Miskin" if mk == 1 else "Tidak Miskin",
            "Rata-rata Pengeluaran Non-Makanan per Kapita (Rp)": round(
                weighted_mean(grp["nonfoodkapita"], grp["weind"]), 0
            ),
        }
    )
result_rows.append(
    {
        "Status Kemiskinan": "Total",
        "Rata-rata Pengeluaran Non-Makanan per Kapita (Rp)": round(
            weighted_mean(df_kp43["nonfoodkapita"], df_kp43["weind"]), 0
        ),
    }
)
df_t55 = pd.DataFrame(result_rows).set_index("Status Kemiskinan")
write_table(
    ws,
    "Tabel 55. Rata-rata Pengeluaran Non-Makanan per Kapita Menurut Status Kemiskinan",
    df_t55,
)

# ============================================================
# TABLE 56: PKH Recipient (r2002) by Poverty
# ============================================================

ws = get_ws("T56_PKH")
result_rows = []
for r2002_val in sorted(df_rt["r2002"].dropna().unique()):
    row = {"Penerima PKH": int(r2002_val)}
    for mk in [0, 1]:
        mask = (df_rt["r2002"] == r2002_val) & (df_rt["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(df_rt.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
df_t56 = pd.DataFrame(result_rows).set_index("Penerima PKH")
for col in df_t56.columns:
    total = df_t56[col].sum()
    df_t56[col.replace("N ", "% ")] = (df_t56[col] / total * 100).round(2)
write_table(ws, "Tabel 56. Penerima PKH Menurut Status Kemiskinan", df_t56)

# ============================================================
# TABLE 57: BPNT/Sembako Recipient (r2005) by Poverty
# ============================================================

ws = get_ws("T57_BPNT")
result_rows = []
for r2005_val in sorted(df_rt["r2005"].dropna().unique()):
    row = {"Penerima BPNT/Sembako": int(r2005_val)}
    for mk in [0, 1]:
        mask = (df_rt["r2005"] == r2005_val) & (df_rt["mkako"] == mk)
        row[f"N mkako={int(mk)}"] = round(df_rt.loc[mask, "fwt"].sum(), 0)
    result_rows.append(row)
df_t57 = pd.DataFrame(result_rows).set_index("Penerima BPNT/Sembako")
for col in df_t57.columns:
    total = df_t57[col].sum()
    df_t57[col.replace("N ", "% ")] = (df_t57[col] / total * 100).round(2)
write_table(ws, "Tabel 57. Penerima BPNT/Sembako Menurut Status Kemiskinan", df_t57)

# ============================================================
# TABLE 58-59: APS with Standard Error (Complex Sample Approximation)
# ============================================================

ws = get_ws("T58_59_APS_SE")
sub = df_ind[(df_ind["r407"] >= 7) & (df_ind["r407"] <= 18)].copy()
result_rows = []
for mk in [0, 1]:
    for age_grp in ["7-12", "13-15", "16-18"]:
        mask = (sub["mkako"] == mk) & (sub["kelum4"].astype(str) == age_grp)
        grp = sub[mask]
        if len(grp) == 0:
            continue
        mn = weighted_mean(grp["aps"], grp["fwt"])
        n = grp["fwt"].sum()
        se = np.sqrt(mn * (100 - mn) / n) if n > 0 and pd.notna(mn) else np.nan
        result_rows.append(
            {
                "Status Kemiskinan": "Miskin" if mk == 1 else "Tidak Miskin",
                "Kelompok Umur": age_grp,
                "APS (%)": round(mn, 2),
                "SE": round(se, 4),
                "CV": round(se / mn * 100, 2) if mn and mn > 0 else np.nan,
            }
        )
df_t5859 = pd.DataFrame(result_rows).set_index(["Status Kemiskinan", "Kelompok Umur"])
write_table(
    ws,
    "Tabel 58-59. APS (7-18 th) dengan Standard Error Menurut Status Kemiskinan",
    df_t5859,
)

# ============================================================
# TABLE 60-61: Employment with SE
# ============================================================

ws = get_ws("T60_61_Employment_SE")
sub = df_ind[df_ind["r407"] >= 15].copy()
result_rows = []
for mk in [0, 1]:
    mask = sub["mkako"] == mk
    grp = sub[mask]
    w = grp["fwt"]
    n = w.sum()
    for var, label in [
        ("tkerja", "Tidak Bekerja"),
        ("ktani", "Pertanian"),
        ("kntani", "Non-Pertanian"),
    ]:
        mn = weighted_mean(grp[var], w)
        se = np.sqrt(mn * (100 - mn) / n) if n > 0 and pd.notna(mn) else np.nan
        result_rows.append(
            {
                "Status Kemiskinan": "Miskin" if mk == 1 else "Tidak Miskin",
                "Indikator Kerja": label,
                "Rata-rata (%)": round(mn, 2),
                "SE": round(se, 4),
                "CV": round(se / mn * 100, 2) if mn and mn > 0 else np.nan,
            }
        )
df_t6061 = pd.DataFrame(result_rows).set_index(["Status Kemiskinan", "Indikator Kerja"])
write_table(
    ws,
    "Tabel 60-61. Status Kerja (Sektor) dengan SE Menurut Status Kemiskinan (15+)",
    df_t6061,
)

# ============================================================
# TABLE 62-63: Formal/Informal Employment with SE
# ============================================================

ws = get_ws("T62_63_FormalInformal_SE")
result_rows = []
for mk in [0, 1]:
    mask = sub["mkako"] == mk
    grp = sub[mask]
    w = grp["fwt"]
    n = w.sum()
    for var, label in [
        ("tkerja", "Tidak Bekerja"),
        ("kformal", "Formal"),
        ("kinformal", "Informal"),
    ]:
        mn = weighted_mean(grp[var], w)
        se = np.sqrt(mn * (100 - mn) / n) if n > 0 and pd.notna(mn) else np.nan
        result_rows.append(
            {
                "Status Kemiskinan": "Miskin" if mk == 1 else "Tidak Miskin",
                "Indikator Kerja": label,
                "Rata-rata (%)": round(mn, 2),
                "SE": round(se, 4),
                "CV": round(se / mn * 100, 2) if mn and mn > 0 else np.nan,
            }
        )
df_t6263 = pd.DataFrame(result_rows).set_index(["Status Kemiskinan", "Indikator Kerja"])
write_table(
    ws,
    "Tabel 62-63. Status Kerja (Formal/Informal) dengan SE Menurut Status Kemiskinan (15+)",
    df_t6263,
)

# ============================================================
# TABLE 64-65: Water & Sanitation with SE
# ============================================================

ws = get_ws("T64_65_WaterSan_SE")
result_rows = []
for mk in [0, 1]:
    mask = df_rt["mkako"] == mk
    grp = df_rt[mask]
    w = grp["fwt"]
    n = w.sum()
    for var, label in [
        ("airmlayak", "Air Minum Layak"),
        ("sab", "Air Bersih (SAB)"),
        ("sal", "Sanitasi Layak"),
    ]:
        mn = weighted_mean(grp[var], w)
        se = np.sqrt(mn * (100 - mn) / n) if n > 0 and pd.notna(mn) else np.nan
        result_rows.append(
            {
                "Status Kemiskinan": "Miskin" if mk == 1 else "Tidak Miskin",
                "Indikator": label,
                "Rata-rata (%)": round(mn, 2),
                "SE": round(se, 4),
                "CV": round(se / mn * 100, 2) if mn and mn > 0 else np.nan,
            }
        )
df_t6465 = pd.DataFrame(result_rows).set_index(["Status Kemiskinan", "Indikator"])
write_table(
    ws,
    "Tabel 64-65. Akses Air dan Sanitasi dengan SE Menurut Status Kemiskinan",
    df_t6465,
)

# ============================================================
# TABLE 66-69: RSE for P0, P1, P2 and Poor Count
# ============================================================

ws = get_ws("T66_69_RSE_Poverty")
result_rows = []
for r102_val, grp in df_kp43.groupby("r102"):
    w = grp["weind"]
    n = w.sum()

    # P0
    mn_p0 = weighted_mean(grp["dmkako"], w)
    se_p0 = np.sqrt(mn_p0 * (100 - mn_p0) / n) if n > 0 and pd.notna(mn_p0) else np.nan

    # P1
    mn_p1 = weighted_mean(grp["p1kako"], w)
    se_p1 = grp["p1kako"].std() / np.sqrt(len(grp)) if len(grp) > 1 else np.nan

    # P2
    mn_p2 = weighted_mean(grp["p2kako"], w)
    se_p2 = grp["p2kako"].std() / np.sqrt(len(grp)) if len(grp) > 1 else np.nan

    # Poor count
    jml = grp.loc[grp["mkako"] == 1, "weind"].sum()

    result_rows.append(
        {
            "Wilayah (r102)": int(r102_val),
            "P0 (%)": round(mn_p0, 2),
            "SE P0": round(se_p0, 4),
            "RSE P0 (%)": round(se_p0 / mn_p0 * 100, 2)
            if mn_p0 and mn_p0 > 0
            else np.nan,
            "P1": round(mn_p1, 3),
            "SE P1": round(se_p1, 4),
            "RSE P1 (%)": round(se_p1 / mn_p1 * 100, 2)
            if mn_p1 and mn_p1 > 0
            else np.nan,
            "P2": round(mn_p2, 3),
            "SE P2": round(se_p2, 4),
            "RSE P2 (%)": round(se_p2 / mn_p2 * 100, 2)
            if mn_p2 and mn_p2 > 0
            else np.nan,
            "Jumlah Miskin": round(jml, 0),
        }
    )
df_t6669 = pd.DataFrame(result_rows).set_index("Wilayah (r102)")
write_table(ws, "Tabel 66-69. RSE P0, P1, P2, dan Jumlah Penduduk Miskin", df_t6669)

# ============================================================
# CHART DATA: Decile x Water/Sanitation Access (Grafik 1-3)
# ============================================================

ws = get_ws("G1_3_ChartData")
result_rows = []
for dec in sorted(df_rt["nkapita"].dropna().unique()):
    grp = df_rt[df_rt["nkapita"] == dec]
    w = grp["fwt"]
    result_rows.append(
        {
            "Desil Pengeluaran": int(dec),
            "Air Minum Layak (%)": round(weighted_mean(grp["airmlayak"], w), 2),
            "Air Bersih/SAB (%)": round(weighted_mean(grp["sab"], w), 2),
            "Sanitasi Layak (%)": round(weighted_mean(grp["sal"], w), 2),
        }
    )
df_chart = pd.DataFrame(result_rows).set_index("Desil Pengeluaran")
write_table(
    ws, "Data Grafik 1-3: Akses Air dan Sanitasi Menurut Desil Pengeluaran", df_chart
)

# ============================================================
# SAVE WORKBOOK
# ============================================================

print(f"Saving to {OUTPUT} ...")
wb.save(OUTPUT)
print("Done! File saved successfully.")
print(f"Total sheets: {len(wb.sheetnames)}")
